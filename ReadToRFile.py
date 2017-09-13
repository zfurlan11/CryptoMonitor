import openpyxl
import datetime
import requests
import bs4
import os
import re
import CoinDict
import pprint
import json
import hashlib
import numpy

CD = CoinDict.dict

#Excel Reader that Takes The file and outputs a list for time and a list of lists for each of the recorded attributes to be analyzed by R

def XRead():
    def UpR():
        os.remove('Rfile.txt')
        file = open('Rfile.txt', 'w')
        file.write('time<-' + 'c(' + ','.join(str(e).replace(' ', '') for e in List[0]) + ')\n\n')
        for i in range(0, 251):
            try:
                file.write(CD[i]['name'].replace(' ', '') + 'v <-' + 'c( ' + ','.join(
                    str(e) for e in List[1][i][1]) + ' )\n\n')
                file.write(CD[i]['name'].replace(' ', '') + 'price <-' + 'c( ' + ','.join(
                    str(e) for e in List[1][i][2]) + ' )\n\n')
                print('Coin: ' + CD[i]['name'] + ' Writing ' + str(List[1][i][3][0]) + ' as first hype value')
                file.write(CD[i]['name'].replace(' ', '') + 'hype <-' + 'c( ' + ','.join(
                    str(e) for e in List[1][i][3]) + ' )\n\n')
            except:
                print(str(i) + 'Coin not Found')
                continue
        print(len(List[1][0][1]))
        print(len(List[1][0][2]))
        print(len(List[1][0][3]))
        file.close()
    def CoinRead():
        Mlist = []
        Elist = []
        x = 3
        for i in range(0, 251):
            try:
                Clist = [CD[i]['name'],[],[],[],[]]
                for t in range(3,len(time) + 3):
                    refRead = out.cell(row = x , column = t)
                    refRead2 = out.cell(row=x + 1, column=t)
                    refRead3 = out.cell(row=x + 2, column=t)
                    #refRead4 = out.cell(row=x + 3, column=t)
                    Clist[1].append(refRead.value)
                    Clist[2].append(refRead2.value)
                    Clist[3].append(refRead3.value)
                    #Clist[4].append(refRead4.value)
                #print(Clist)
                #print(CD[i]['name'])
                #print(str(refRead3.value) + 'writing From: ' + str(t) + ' , ' + str(x + 2))
                Mlist.append(Clist)
                x+=5
            except:
                Mlist.append(Elist)
                continue
        #print(Mlist)
        return Mlist
    emptychk = 3
    wb = openpyxl.load_workbook('tester.xlsx')
    out = wb.get_active_sheet()
    #Make Time List
    time = []
    test = []
    testx = 0
    refT = out.cell(row=1, column=emptychk)
    while refT.value != None:
        test.append(testx)
        testx += 1
        time.append(refT.value)
        emptychk += 1
        refT = out.cell(row=1, column=emptychk)
    #print (time)
    List = [test, CoinRead()]
    UpR()
XRead()


