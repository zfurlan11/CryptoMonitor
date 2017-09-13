import requests
import bs4
import os
import re
import CoinDict
import pprint
import json
import hashlib
import openpyxl
import datetime
import time

from selenium import webdriver

CD = CoinDict.dict

def ThreadIDGet():
    os.remove('tids.txt')
    url = 'http://boards.4chan.org/biz/catalog'
    driver = webdriver.Firefox()
    driver.get(url)
    res = driver.page_source
    soup = bs4.BeautifulSoup(res, 'lxml')
    body = soup.find('body')
    content = body.find(id="content")
    thread = content.find(id="threads")
    ThR = re.compile(r'(thread-)(\d\d\d\d\d\d\d)')

    threadlist = ThR.findall(str(thread))
    ThreadIDs = open("tids.txt", 'a')
    for i in range(1, len(threadlist)):
        ThreadIDs.write(threadlist[i][1] + '\n')
    print('IDs Obtained')
    driver.close()
    ThreadIDs.close()

def TextGet():
    os.remove('text.txt')
    tids = open('tids.txt','r')
    tlist = tids.readlines()

    for i in range(0, len(tlist)-1):
        url = 'http://boards.4chan.org/biz/thread/' + tlist[i]
        #print(url)
        res = requests.get(url)

        soup = bs4.BeautifulSoup(res.content,"html.parser")

        textelems = soup.select('blockquote')
        file = open('text.txt','a')

        idf = re.compile(r'(id="m)(\d\d\d\d\d\d\d)')

        file.write('Thread#' + str(i) + '\n')
        for i in range(0,len(textelems)):
            mo = idf.search(str(textelems[i]))
            id = mo.group(2)
            dat = str(textelems[i].getText().encode('utf-8'))+'\n'
            file.write(str(id) + '::' + dat[1:])
        file.write('\n')
        file.close()
    print('Scrape Complete')

def Count(coin):
    file = open('text.txt', 'r')
    postlist = file.readlines()
    checklist = [CD[coin]['aka'][0],CD[coin]['aka'][0].lower(),(CD[coin]['name']),(CD[coin]['name'].lower())]
    Count = 0
    print(checklist)
    for i in range(0,len(postlist)):
        for x in checklist:
            if x in postlist[i].split():
                Count += 1
                break
            else:
                continue
    file.close()
    return Count

def NewBook(Name):
    wb = openpyxl.Workbook()
    print(type(wb))
    inp = wb.get_active_sheet()
    print(type(inp))
    inp['B1']='TIME'
    inp['A2']='Coin'
    inp['B2'] = 'Metric'
    inp.freeze_panes = 'C1'

    x = 0
    for i in range(0, 251):
        try:
            inp['A{}'.format(str(x + 3))] = CD[i]['name']
            inp['B{}'.format(str(x + 3))] = 'BVol'
            inp['B{}'.format(str(x + 4))] = 'Price'
            inp['B{}'.format(str(x + 5))] = 'Posts'
            x += 5
        except:
            continue
    #for i in range(0,90):


    wb.save(Name + '.xlsx')

def Update():
    emptychk = 3
    x = 3
    wb = openpyxl.load_workbook('tester.xlsx')
    inp = wb.get_active_sheet()
    refC = inp.cell(row = 3,column = emptychk)
    while refC.value != None:
        emptychk += 1
        refC = inp.cell(row=3, column=emptychk)
    inp.cell(row = 1, column = emptychk).value = datetime.datetime.now()
    #Update Volume
    print('UPDATING')
    for i in range(0, 251):
        try:
            url = 'https://bittrex.com/api/v1.1/public/getmarketsummary?market=btc-' + CD[i]['aka'][0].lower()
            data = json.loads(requests.get(url).text)
            data = data['result']
            volume = data[0]['BaseVolume']
            price = data[0]['Last']
            VCell = inp.cell(row=x, column=emptychk)
            PCell = inp.cell(row=x + 1, column=emptychk)
            CCell = inp.cell(row=x + 2, column=emptychk)
            VCell.value = volume
            PCell.value = price
            CCell.value = Count(i)
            print(str(i))
            print(str(CCell.value) + 'writing to: ' + str(emptychk) + ' , ' + str(x + 2))
            x += 5
        except:
            print(str(i) + ' Coin Not Found')
            continue
    print('UPDATED')
    wb.save('tester.xlsx')

NewBook('tester')
while True:
    ThreadIDGet()
    TextGet()
    Update()


